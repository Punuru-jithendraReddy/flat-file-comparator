import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import time
import textwrap

# --- 1. CONFIGURATION & SETUP ---
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

st.set_page_config(
    page_title="Flat File Comparison Tool", 
    page_icon="📊",
    layout="wide"
)

# --- 2. CSS STYLING (Excel Replica Theme) ---
st.markdown("""
<style>
    /* 1. GREEN BUTTON (Strict) */
    div.stButton > button {
        background-color: #28a745 !important; 
        color: white !important;
        border-color: #28a745 !important;
        font-weight: bold !important;
        width: 100%;
        height: 50px;
        font-size: 18px !important;
    }
    div.stButton > button:hover {
        background-color: #218838 !important;
        border-color: #1e7e34 !important;
    }

    /* 2. BLUE MULTISELECT TAGS */
    span[data-baseweb="tag"] {
        background-color: #4472C4 !important; /* Excel Blue */
        color: white !important;
    }

    /* 3. REPORT DASHBOARD STYLES (HTML/CSS) */
    .report-container {
        font-family: 'Calibri', sans-serif;
        border: 1px solid #dcdcdc;
        margin-bottom: 20px;
        background-color: white;
    }
    .report-header {
        background-color: #4472C4; /* Excel Blue Header */
        color: white;
        padding: 8px 15px;
        font-weight: bold;
        font-size: 16px;
        border-bottom: 1px solid #bbb;
    }
    .report-row {
        display: flex;
        border-bottom: 1px solid #eee;
    }
    .report-row:last-child {
        border-bottom: none;
    }
    .report-key {
        width: 30%;
        background-color: #f2f2f2;
        padding: 8px 15px;
        font-weight: 600;
        color: #333;
        border-right: 1px solid #eee;
    }
    .report-val {
        width: 70%;
        padding: 8px 15px;
        color: #000;
    }
    .status-good { color: green; font-weight: bold; }
    .status-bad { color: #d9534f; font-weight: bold; }
    .status-neutral { color: #f0ad4e; font-weight: bold; }
    
    /* RECO BOX */
    .reco-box {
        background-color: #d4edda;
        color: #155724;
        padding: 10px;
        border-left: 5px solid #28a745;
        margin: 10px;
        font-weight: bold;
    }

    /* FOOTER */
    .footer {
        position: fixed;
        left: 0; bottom: 0; width: 100%;
        background-color: #1a202c;
        color: white;
        text-align: center;
        padding: 10px;
        font-size: 14px;
        z-index: 99999;
    }
    .footer a { color: #63b3ed; text-decoration: none; font-weight: bold; }
    .block-container { padding-bottom: 80px; }
</style>
""", unsafe_allow_html=True)

# --- 3. LOGIC FUNCTIONS ---

def normalize_for_comparison(series, is_case_insensitive_data, should_trim_whitespace):
    s_numeric = pd.to_numeric(series, errors='coerce')
    s = series.where(s_numeric.isna(), s_numeric)
    s = pd.to_datetime(s, errors='coerce').dt.strftime('%Y-%m-%d').fillna(s)
    s = s.astype(str)
    s = s.str.replace(r'\.0$', '', regex=True)
    s_lower_for_nulls = s.str.lower().str.strip()
    s[s_lower_for_nulls.isin(['nan', '<na>', 'none', 'nat', ''])] = ''
    
    if should_trim_whitespace:
        s = s.str.strip().str.replace(r'\s+', ' ', regex=True) 
    if is_case_insensitive_data:
        s = s.str.lower()
    return s

def smart_read_file(file_obj, header_row, sheet_name=0):
    file_ext = file_obj.name.split('.')[-1].lower()
    try:
        if file_ext in ['xlsx', 'xls']:
            return pd.read_excel(file_obj, header=header_row, sheet_name=sheet_name)
        elif file_ext == 'csv':
            file_obj.seek(0)
            try: return pd.read_csv(file_obj, encoding='utf-8-sig', header=header_row)
            except: 
                file_obj.seek(0)
                return pd.read_csv(file_obj, encoding='latin1', header=header_row)
    except Exception as e:
        st.error(f"Error: {e}")
        return None

def get_diagnosis(pct):
    if pct == 100: return "Files are identical."
    elif pct >= 95: return "High Accuracy (Minor Differences)"
    elif pct > 0: return "Mismatch Found"
    else: return "Critical Mismatch (No rows matched)"

# --- 4. SIDEBAR ---
with st.sidebar:
    st.image("https://cdn-icons-png.flaticon.com/512/10891/10891404.png", width=70)
    st.title("Settings")
    
    with st.expander("🛠 Options", expanded=True):
        opt_case_cols = st.checkbox(
            "Case-Insensitive Cols", 
            value=True,
            help="ℹ️ If selected, 'ColumnName' and 'columnname' will be matched as the same column."
        )
        opt_case_data = st.checkbox(
            "Case-Insensitive Data", 
            value=True,
            help="ℹ️ If selected, text data like 'Apple' and 'apple' will be considered identical matches."
        )
        opt_trim = st.checkbox(
            "Trim Whitespace", 
            value=True,
            help="ℹ️ If selected, extra spaces at the beginning or end of text (e.g., '  Data ') will be removed before comparing."
        )
    
    with st.expander("📑 Excel Output", expanded=True):
        gen_row_sheet = st.checkbox("Row Comparison", value=True, help="ℹ️ Include a sheet showing exactly which rows are missing or added.")
        gen_col_sheet = st.checkbox("Column Analysis", value=True, help="ℹ️ Include a sheet comparing the list of column headers.")
        gen_uniq_sheet = st.checkbox("Unique Values", value=True, help="ℹ️ Include a sheet listing unique values found in your key columns.")
        gen_stats_sheet = st.checkbox("Summary Stats", value=True, help="ℹ️ Include a sheet with Count, Sum, Min, and Max for numeric columns.")

    st.markdown("---")
    st.markdown("### 👨‍💻 Developer")
    st.markdown("**Jithendra Reddy**")
    st.markdown("📧 [Email](mailto:jithendrareddypunuru@gmail.com)")
    st.markdown("🔗 [LinkedIn](https://www.linkedin.com/in/jithendrareddypunuru/)")

# --- 5. MAIN UI ---

st.title("📂 Flat File Comparison Tool")

with st.expander("ℹ️ Click here for Instructions"):
    st.markdown("""
    **How to use this tool:**
    1.  **Upload Files:** Drag and drop your **Source** (Old) and **Target** (New) files.
    2.  **Select Sheet:** If your Excel file has multiple sheets, a dropdown will appear.
    3.  **Check Headers:** Ensure the 'Header Row' number matches where your column titles are.
    4.  **Select Keys:** Once uploaded, select the column(s) that identify a unique row (e.g., Employee ID).
    5.  **Run:** Click the green button to generate the report.
    """)

st.markdown("Upload two files below to generate a detailed comparison report.")

col_input1, col_input2 = st.columns(2)

# --- SOURCE FILE INPUT ---
with col_input1:
    st.subheader("Source File")
    src_file = st.file_uploader(
        "Upload Source", 
        type=["xlsx", "xls", "csv"], 
        key="src",
        help="ℹ️ Drag and drop your Original / Baseline file here."
    )
    
    # Dynamic Sheet Selection Logic
    src_sheet_name = 0
    if src_file:
        file_ext = src_file.name.split('.')[-1].lower()
        if file_ext in ['xlsx', 'xls']:
            try:
                # Read specific sheet names efficiently
                xl_file = pd.ExcelFile(src_file)
                sheet_names = xl_file.sheet_names
                src_file.seek(0) # IMPORTANT: Reset pointer after reading names
                
                if len(sheet_names) > 1:
                    src_sheet_name = st.selectbox(
                        "📄 Select Source Sheet", 
                        options=sheet_names, 
                        key="src_sheet_select"
                    )
                else:
                    src_sheet_name = sheet_names[0]
                    st.info(f"Using sheet: {src_sheet_name}")
            except Exception as e:
                st.error(f"Error reading sheets: {e}")

    src_header = st.number_input(
        "Header Row (Source)", 
        min_value=1, 
        value=1, 
        key="h1",
        help="Row number for column headers."
    ) - 1

# --- TARGET FILE INPUT ---
with col_input2:
    st.subheader("Target File")
    tgt_file = st.file_uploader(
        "Upload Target", 
        type=["xlsx", "xls", "csv"], 
        key="tgt",
        help="ℹ️ Drag and drop your New / Updated file here."
    )
    
    # Dynamic Sheet Selection Logic
    tgt_sheet_name = 0
    if tgt_file:
        file_ext = tgt_file.name.split('.')[-1].lower()
        if file_ext in ['xlsx', 'xls']:
            try:
                xl_file = pd.ExcelFile(tgt_file)
                sheet_names = xl_file.sheet_names
                tgt_file.seek(0) # IMPORTANT: Reset pointer after reading names
                
                if len(sheet_names) > 1:
                    tgt_sheet_name = st.selectbox(
                        "📄 Select Target Sheet", 
                        options=sheet_names, 
                        key="tgt_sheet_select"
                    )
                else:
                    tgt_sheet_name = sheet_names[0]
                    st.info(f"Using sheet: {tgt_sheet_name}")
            except Exception as e:
                st.error(f"Error reading sheets: {e}")

    tgt_header = st.number_input(
        "Header Row (Target)", 
        min_value=1, 
        value=1, 
        key="h2",
        help="Row number for column headers."
    ) - 1

# B. Execution
if src_file and tgt_file:
    st.divider()
    
    # Load DataFrames using selected sheets
    df1 = smart_read_file(src_file, src_header, src_sheet_name)
    df2 = smart_read_file(tgt_file, tgt_header, tgt_sheet_name)

    if df1 is not None and df2 is not None:
        total_src_rows = len(df1)
        total_tgt_rows = len(df2)

        # Map Columns
        src_cols = df1.columns
        tgt_cols = df2.columns
        common_cols_list = []
        src_to_tgt_map = {}

        if opt_case_cols:
            src_map = {str(c).lower(): c for c in src_cols}
            tgt_map = {str(c).lower(): c for c in tgt_cols}
            common_lower = set(src_map.keys()) & set(tgt_map.keys())
            for k in common_lower:
                common_cols_list.append(src_map[k])
                src_to_tgt_map[src_map[k]] = tgt_map[k]
        else:
            common = set(src_cols) & set(tgt_cols)
            common_cols_list = list(common)
            src_to_tgt_map = {c: c for c in common}

        if not common_cols_list:
            st.error("❌ No common columns found.")
        else:
            c_sel, c_btn = st.columns([3, 1])
            with c_sel:
                all_options = sorted(common_cols_list, key=str)
                selected_src = st.multiselect(
                    "Select Key Columns (Unique Identifiers)", 
                    options=all_options,
                    default=all_options,
                    help="ℹ️ IMPORTANT: Select the columns that make a row unique (e.g., 'Order ID' or 'Email')."
                )

            with c_btn:
                st.write("") 
                st.write("") 
                run_btn = st.button(
                    "🚀 Run Comparison",
                    help="ℹ️ Click to start the matching process."
                )

            if run_btn:
                if not selected_src:
                    st.error("Select at least one column.")
                else:
                    with st.spinner("Analyzing..."):
                        
                        # 1. PRIMARY MERGE
                        selected_tgt = [src_to_tgt_map[c] for c in selected_src]
                        df1_n = df1[selected_src].copy()
                        df2_n = df2[selected_tgt].copy()
                        df2_n.columns = selected_src

                        for c in selected_src:
                            df1_n[c] = normalize_for_comparison(df1_n[c], opt_case_data, opt_trim)
                            df2_n[c] = normalize_for_comparison(df2_n[c], opt_case_data, opt_trim)

                        df1_n['_oid_src'] = df1.index
                        df2_n['_oid_tgt'] = df2.index

                        merged = pd.merge(df1_n, df2_n, on=selected_src, how='outer', indicator=True)
                        
                        # Primary Stats
                        in_both_idxs = merged[merged['_merge']=='both']
                        c_both = len(in_both_idxs)
                        c_src = len(merged[merged['_merge']=='left_only'])
                        c_tgt = len(merged[merged['_merge']=='right_only'])
                        total_union = c_both + c_src + c_tgt
                        match_pct = (c_both/total_union*100) if total_union else 0
                        diagnosis = get_diagnosis(match_pct)
                        diag_class = "status-good" if match_pct == 100 else ("status-bad" if match_pct < 80 else "status-neutral")

                        # --- SMART RECOVERY & MISMATCH DIAGNOSIS ---
                        mismatch_html = ""
                        reco_msg = ""
                        
                        if match_pct < 100 and len(selected_src) > 1:
                            best_alt_pct = 0
                            best_alt_col = None
                            
                            for col_to_remove in selected_src:
                                temp_keys = [k for k in selected_src if k != col_to_remove]
                                temp_merged = pd.merge(df1_n, df2_n, on=temp_keys, how='inner')
                                temp_match_count = temp_merged['_oid_src'].nunique()
                                temp_union = total_src_rows + total_tgt_rows - temp_match_count
                                temp_pct = (temp_match_count / temp_union * 100) if temp_union else 0
                                
                                if temp_pct > match_pct:
                                    if temp_pct > best_alt_pct:
                                        best_alt_pct = temp_pct
                                        best_alt_col = col_to_remove

                            if best_alt_col:
                                reco_msg = f"""<div class="report-row" style="background-color: #d4edda; border-bottom: 1px solid #c3e6cb;">
<div class="report-key" style="color: #155724; background-color: #d4edda;">💡 Recommendation</div>
<div class="report-val" style="color: #155724;">
Removing the column <b>'{best_alt_col}'</b> from your Key selection would increase the Match Percentage from {match_pct:.2f}% to <strong>{best_alt_pct:.2f}%</strong>.
</div></div>"""

                        if match_pct == 100:
                            mismatch_html = '<div class="report-row"><div class="report-key">Details</div><div class="report-val status-good">No value mismatches found.</div></div>'
                        elif match_pct == 0:
                            if reco_msg:
                                mismatch_html = reco_msg
                            else:
                                mismatch_html = """
                                <div class="report-row">
                                    <div class="report-key">Critical Mismatch Diagnosis</div>
                                    <div class="report-val" style="color:#d9534f">
                                        ⚠️ <strong>No rows matched.</strong><br>
                                        See sample values below to spot formatting differences.
                                    </div>
                                </div>
                                """
                                for col in selected_src:
                                    s_samp = ", ".join(df1_n[col].unique()[:3].astype(str))
                                    t_samp = ", ".join(df2_n[col].unique()[:3].astype(str))
                                    mismatch_html += f"""
                                    <div class="report-row" style="background-color:#fff3cd">
                                        <div class="report-key">Column: {col}</div>
                                        <div class="report-val">
                                            <strong>Source:</strong> {s_samp}... <br>
                                            <strong>Target:</strong> {t_samp}...
                                        </div>
                                    </div>
                                    """
                        else:
                            if reco_msg: mismatch_html = reco_msg
                            value_cols = [c for c in common_cols_list if c not in selected_src]
                            if value_cols and not in_both_idxs.empty:
                                idx_src = in_both_idxs['_oid_src'].astype(int)
                                idx_tgt = in_both_idxs['_oid_tgt'].astype(int)
                                v_df1 = df1.loc[idx_src, value_cols].reset_index(drop=True)
                                v_df2 = df2.loc[idx_tgt, value_cols].reset_index(drop=True)
                                v_df2.columns = [c if c in value_cols else src_to_tgt_map.get(c,c) for c in v_df2.columns] 
                                v_df2 = v_df2[value_cols]
                                
                                mm_counts = []
                                for col in value_cols:
                                    s1 = normalize_for_comparison(v_df1[col], opt_case_data, opt_trim)
                                    s2 = normalize_for_comparison(v_df2[col], opt_case_data, opt_trim)
                                    diff = (s1 != s2).sum()
                                    if diff > 0: mm_counts.append({'Column': col, 'Mismatch Count': diff})
                                
                                mismatch_df = pd.DataFrame(mm_counts)
                                if not mismatch_df.empty:
                                    mismatch_df = mismatch_df.sort_values(by='Mismatch Count', ascending=False)
                                    rows_html = ""
                                    for _, r in mismatch_df.iterrows():
                                        rows_html += f'<div class="report-row"><div class="report-key">{r["Column"]}</div><div class="report-val">{r["Mismatch Count"]:,} rows differ</div></div>'
                                    mismatch_html += rows_html
                                else:
                                    mismatch_html += '<div class="report-row"><div class="report-key">Details</div><div class="report-val status-good">No value mismatches found in matched rows.</div></div>'

                        # --- DISPLAY REPORT ---
                        st.success("✅ Analysis Complete")
                        
                        st.markdown(f"""
                        <div class="report-container">
                            <div class="report-header">File Information</div>
                            <div class="report-row"><div class="report-key">Source File</div><div class="report-val">{src_file.name} (Sheet: {src_sheet_name})</div></div>
                            <div class="report-row"><div class="report-key">Source Rows</div><div class="report-val">{total_src_rows:,}</div></div>
                            <div class="report-row"><div class="report-key">Target File</div><div class="report-val">{tgt_file.name} (Sheet: {tgt_sheet_name})</div></div>
                            <div class="report-row"><div class="report-key">Target Rows</div><div class="report-val">{total_tgt_rows:,}</div></div>
                        </div>
                        """, unsafe_allow_html=True)

                        st.markdown(f"""
                        <div class="report-container">
                            <div class="report-header">Comparison Statistics</div>
                            <div class="report-row"><div class="report-key">Match Percentage</div><div class="report-val" style="font-weight:bold">{match_pct:.2f}%</div></div>
                            <div class="report-row"><div class="report-key">Matched Rows</div><div class="report-val">{c_both:,}</div></div>
                            <div class="report-row"><div class="report-key">Missing in Target</div><div class="report-val">{c_src:,}</div></div>
                            <div class="report-row"><div class="report-key">New in Target</div><div class="report-val">{c_tgt:,}</div></div>
                        </div>
                        """, unsafe_allow_html=True)

                        st.markdown(f"""
                        <div class="report-container">
                            <div class="report-header">Mismatch Diagnosis</div>
                            <div class="report-row"><div class="report-key">Status</div><div class="report-val {diag_class}">{diagnosis}</div></div>
                            {mismatch_html}
                        </div>
                        """, unsafe_allow_html=True)

                        # --- EXCEL GENERATION ---
                        only_src = df1.loc[merged[merged['_merge']=='left_only']['_oid_src'].dropna()].reindex(columns=selected_src)
                        only_tgt = df2.loc[merged[merged['_merge']=='right_only']['_oid_tgt'].dropna()].reindex(columns=selected_tgt)
                        in_both  = df1.loc[in_both_idxs['_oid_src'].dropna()].reindex(columns=selected_src)

                        buffer = BytesIO()
                        wb = Workbook()
                        wb.remove(wb.active)

                        header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        
                        def write_section(ws, r, title):
                            cell = ws.cell(row=r, column=1, value=title)
                            cell.font = header_font; cell.fill = header_fill
                            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                            return r + 1

                        def write_pair(ws, r, k, v):
                            ws.cell(r,1,k).font = Font(bold=True)
                            ws.cell(r,2,v).alignment = Alignment(horizontal='left')
                            return r + 1

                        ws_sum = wb.create_sheet("Executive Summary")
                        ws_sum.column_dimensions['A'].width = 35; ws_sum.column_dimensions['B'].width = 80
                        row = 1
                        row = write_section(ws_sum, row, "File Information")
                        row = write_pair(ws_sum, row, "Source", f"{src_file.name} (Sheet: {src_sheet_name})")
                        row = write_pair(ws_sum, row, "Target", f"{tgt_file.name} (Sheet: {tgt_sheet_name})")
                        row += 1
                        row = write_section(ws_sum, row, "Stats")
                        row = write_pair(ws_sum, row, "Matched", f"{c_both:,}")
                        row = write_pair(ws_sum, row, "Match %", f"{match_pct:.2f}%")

                        if gen_col_sheet:
                            ws = wb.create_sheet("Column Names")
                            ws.append(["Column Name", "In Source", "In Target"])
                            for c in sorted(list(set(df1.columns) | set(df2.columns)), key=str):
                                ws.append([c, "Yes" if c in df1.columns else "No", "Yes" if c in df2.columns else "No"])

                        if gen_row_sheet:
                            ws = wb.create_sheet("Row Comparison")
                            ws.append(['Status'] + selected_src)
                            for _, r in only_src.iterrows(): ws.append(['Only in Source'] + r.astype(str).tolist())
                            for _, r in only_tgt.iterrows(): ws.append(['Only in Target'] + r.astype(str).tolist())
                            for _, r in in_both.iterrows(): ws.append(['In Both'] + r.astype(str).tolist())

                        if gen_uniq_sheet:
                            ws = wb.create_sheet("Unique Values")
                            col_idx = 1
                            for c in selected_src:
                                ws.cell(1, col_idx, c).font = Font(bold=True)
                                s_v = set(df1_n[c].dropna()[df1_n[c]!='']); t_v = set(df2_n[c].dropna()[df2_n[c]!=''])
                                for i, v in enumerate(sorted(s_v - t_v), 3): ws.cell(i, col_idx, v)
                                for i, v in enumerate(sorted(t_v - s_v), 3): ws.cell(i, col_idx+1, v)
                                col_idx += 3

                        if gen_stats_sheet:
                            ws = wb.create_sheet("Summary Stats")
                            nums = [c for c in selected_src if pd.api.types.is_numeric_dtype(df1[c])]
                            col_idx = 1
                            for c in nums:
                                tgt_c = src_to_tgt_map.get(c,c)
                                ws.cell(1, col_idx, c).font = Font(bold=True)
                                for i, s in enumerate(['count','sum','mean','min','max'], 3):
                                    ws.cell(i, col_idx, s)
                                    try: ws.cell(i, col_idx+1, getattr(df1[c], s)())
                                    except: pass
                                    try: ws.cell(i, col_idx+2, getattr(df2[tgt_c], s)())
                                    except: pass
                                col_idx += 4

                        wb.save(buffer)
                        buffer.seek(0)
                        
                        st.download_button(
                            label="📥 Download Full Report (Excel)",
                            data=buffer,
                            file_name=f"Comparison_Report_{datetime.datetime.now().strftime('%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

# --- 6. FOOTER ---
st.markdown("""
<div class="footer">
    Developed by <strong>Jithendra Reddy</strong> | 
    <a href="mailto:jithendrareddypunuru@gmail.com">jithendrareddypunuru@gmail.com</a> | 
    <a href="https://www.linkedin.com/in/jithendrareddypunuru/" target="_blank">LinkedIn Profile</a>
</div>
""", unsafe_allow_html=True)